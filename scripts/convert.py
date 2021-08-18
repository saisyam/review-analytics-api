from openpyxl import load_workbook
import sys
import csv

def excel_to_csv(infile, outfile):
    fields = ['Review', 'Food', 'Service', 'Ambience', 'Price', 'Drinks'] 
    wb = load_workbook(infile)
    out_file = open(outfile, "w")
    csvwriter = csv.writer(out_file)
    csvwriter.writerow(fields)
    sheet = wb.worksheets[0]
    first_row = 1
    for row in sheet.iter_rows():
        if first_row == 1:
            first_row = 0
            continue
        csv_row = []
        csv_row.append(row[3].value)
        for i in range(0, 5):
            csv_row.append(0)
        labels = row[4].value.split(",")
        for label in labels:
            if "food" in label:
                csv_row[1] = 1
            elif "service" in label:
                csv_row[2] = 1
            elif "ambience" in label:
                csv_row[3] = 1
            elif "price" in label:
                csv_row[4] = 1
            elif "drinks" in label:
                csv_row[5] = 1
        csvwriter.writerow(csv_row)
    out_file.close()    


if len(sys.argv) == 3:
    excel_to_csv(sys.argv[1], sys.argv[2])
else:
    print("Usage: python3 convert.py <input file> <output file>")
    exit()
